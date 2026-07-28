#!/usr/bin/env python3
"""Génère un classeur Excel de cadrage de formation à partir d'un JSON.

Usage: python generate_cadrage_xlsx.py input.json output.xlsx
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
DARK = "1F3864"
LIGHT = "D9E2F3"
RED = "C00000"
GREY = "808080"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def title_block(ws, data, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=data.get("titre", "Cadrage formation"))
    t.font = Font(name=FONT, bold=True, size=14, color=DARK)
    sub = []
    if data.get("client"):
        sub.append(f"Client : {data['client']}")
    if data.get("formation"):
        sub.append(f"Formation : {data['formation']}")
    if data.get("date_formation"):
        sub.append(f"Date formation : {data['date_formation']}")
    if data.get("date_cadrage"):
        sub.append(f"Appel de cadrage : {data['date_cadrage']}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value="  |  ".join(sub))
    s.font = Font(name=FONT, size=10, color=GREY)


def build_questions(wb, data):
    ws = wb.active
    ws.title = "Questions de cadrage"
    headers = ["Thème", "Question", "Priorité", "Pourquoi / contexte",
               "Réponse client", "Statut"]
    widths = [32, 62, 16, 48, 48, 14]
    title_block(ws, data, len(headers))
    header_row(ws, 4, headers, widths)

    dv = DataValidation(type="list", formula1='"À poser,Posée,Répondu,Non pertinent"',
                        allow_blank=True)
    ws.add_data_validation(dv)

    r = 5
    for q in data.get("questions", []):
        vals = [q.get("theme", ""), q.get("question", ""),
                q.get("priorite", ""), q.get("pourquoi", ""), "", "À poser"]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        p = ws.cell(row=r, column=3)
        p.alignment = Alignment(horizontal="center", vertical="top")
        if str(p.value).strip().upper().startswith("INDISP"):
            p.font = Font(name=FONT, size=10, bold=True, color=RED)
        else:
            p.font = Font(name=FONT, size=10, color=GREY)
        dv.add(ws.cell(row=r, column=6))
        r += 1

    ws.auto_filter.ref = f"A4:F{max(r - 1, 4)}"
    ws.freeze_panes = "A5"


def build_participants(wb, data):
    parts = data.get("participants")
    if not parts:
        return
    ws = wb.create_sheet("Participants")
    headers = ["Nom", "Prénom", "Rôle identifié", "Entité", "Profil", "Statut"]
    widths = [18, 16, 42, 30, 30, 16]
    header_row(ws, 1, headers, widths)
    for r, p in enumerate(parts, start=2):
        vals = [p.get("nom", ""), p.get("prenom", ""), p.get("role", ""),
                p.get("entite", ""), p.get("profil", ""), p.get("statut", "")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        if "confirm" in str(p.get("statut", "")).lower():
            ws.cell(row=r, column=6).font = Font(name=FONT, size=10, bold=True, color=RED)
    ws.auto_filter.ref = f"A1:F{len(parts) + 1}"
    ws.freeze_panes = "A2"


def build_contexte(wb, data):
    ctx, vig = data.get("contexte"), data.get("points_vigilance")
    if not ctx and not vig:
        return
    ws = wb.create_sheet("Contexte")
    ws.column_dimensions["A"].width = 110
    r = 1
    for label, items in (("Contexte", ctx), ("Points de vigilance", vig)):
        if not items:
            continue
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=DARK)
        r += 1
        for it in items:
            c = ws.cell(row=r, column=1, value=f"• {it}")
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            r += 1
        r += 1


def main():
    if len(sys.argv) != 3:
        sys.exit("Usage: python generate_cadrage_xlsx.py input.json output.xlsx")
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)
    wb = Workbook()
    build_questions(wb, data)
    build_participants(wb, data)
    build_contexte(wb, data)
    wb.save(sys.argv[2])
    print(f"OK: {sys.argv[2]}")


if __name__ == "__main__":
    main()
