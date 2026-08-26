#!/usr/bin/env python3
"""Génère un classeur Excel d'exigences d'appel d'offres à partir d'un JSON.

Usage: python generate_exigences_xlsx.py input.json output.xlsx
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
DARK = "1F3864"
LIGHT = "D9E2F3"
RED = "C00000"
GREY = "808080"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def title_block(ws, data, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=data.get("titre", "Analyse appel d'offres"))
    t.font = Font(name=FONT, bold=True, size=14, color=DARK)
    sub = []
    if data.get("client"):
        sub.append(f"Client : {data['client']}")
    if data.get("objet"):
        sub.append(f"Objet : {data['objet']}")
    if data.get("date_remise"):
        sub.append(f"Remise offre : {data['date_remise']}")
    if data.get("type_dossier"):
        sub.append(f"Type de dossier : {data['type_dossier']}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value="  |  ".join(sub))
    s.font = Font(name=FONT, size=10, color=GREY)


def build_exigences(wb, data):
    ws = wb.active
    ws.title = "Exigences CCTP"
    headers = ["N°", "Source", "Thème", "Exigence", "Catégorie",
               "Critère de notation lié", "Statut de traitement",
               "Partie du mémoire", "Page de réponse", "Commentaire / risque"]
    widths = [8, 20, 22, 50, 14, 30, 16, 26, 12, 40]
    title_block(ws, data, len(headers))
    header_row(ws, 4, headers, widths)

    dv_statut = DataValidation(
        type="list",
        formula1='"Non traité,En cours,Traité,Non applicable (à justifier)"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_statut)

    r = 5
    for ex in data.get("exigences", []):
        vals = [
            ex.get("numero", ""), ex.get("source", ""), ex.get("theme", ""),
            ex.get("exigence", ""), ex.get("categorie", ""),
            ex.get("critere_notation", ""),
            ex.get("statut", "Non traité"),
            ex.get("partie_memoire", ""), ex.get("page_reponse", ""),
            ex.get("commentaire", ""),
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        cat = ws.cell(row=r, column=5)
        cat.alignment = Alignment(horizontal="center", vertical="top")
        categorie = str(cat.value).strip().upper()
        if categorie == "ÉLIMINATOIRE" or categorie == "ELIMINATOIRE":
            cat.font = Font(name=FONT, size=10, bold=True, color=RED)
        elif categorie == "OBLIGATOIRE":
            cat.font = Font(name=FONT, size=10, bold=True, color=DARK)
        else:
            cat.font = Font(name=FONT, size=10, color=GREY)
        dv_statut.add(ws.cell(row=r, column=7))
        r += 1

    ws.auto_filter.ref = f"A4:J{max(r - 1, 4)}"
    ws.freeze_panes = "A5"


def build_jalons(wb, data):
    jalons = data.get("jalons")
    if not jalons:
        return
    ws = wb.create_sheet("Deadline & jalons")
    headers = ["Jalon", "Date", "Contrainte associée"]
    widths = [30, 24, 60]
    header_row(ws, 1, headers, widths)
    for r, j in enumerate(jalons, start=2):
        vals = [j.get("jalon", ""), j.get("date", ""), j.get("contrainte", "")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
    ws.auto_filter.ref = f"A1:C{len(jalons) + 1}"
    ws.freeze_panes = "A2"


def build_synthese_texte(wb, title, items):
    if not items:
        return
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 110
    for r, it in enumerate(items, start=1):
        c = ws.cell(row=r, column=1, value=f"• {it}")
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)


def build_texte_libre(wb, title, text):
    """Onglet à une seule cellule de texte libre, sans puce — pour un contenu
    unique potentiellement long (ex. la description d'un format de réponse
    imposé), à la différence de build_synthese_texte qui liste plusieurs
    points courts."""
    if not text:
        return
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 110
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 80


def build_personnes(wb, data):
    personnes = data.get("personnes")
    if not personnes:
        return
    ws = wb.create_sheet("Personnes liées à l'AO")
    headers = ["Nom", "Prénom", "Rôle dans l'AO", "Poste actuel", "Séniorité",
               "Profil", "Statut"]
    widths = [18, 16, 32, 30, 18, 24, 16]
    header_row(ws, 1, headers, widths)
    for r, p in enumerate(personnes, start=2):
        vals = [
            p.get("nom", ""), p.get("prenom", ""), p.get("role_ao", ""),
            p.get("poste_actuel", ""), p.get("seniorite", ""),
            p.get("profil", ""), p.get("statut", ""),
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        if "confirm" in str(p.get("statut", "")).lower():
            ws.cell(row=r, column=7).font = Font(name=FONT, size=10, bold=True, color=RED)
    ws.auto_filter.ref = f"A1:G{len(personnes) + 1}"
    ws.freeze_panes = "A2"


def build_technologies(wb, data):
    technos = data.get("technologies")
    if not technos:
        return
    ws = wb.create_sheet("Technologies mentionnées")
    headers = ["Techno / Méthodologie", "Citée où", "État de l'art",
               "Maturité", "Alternatives", "Point de vigilance"]
    widths = [26, 22, 40, 14, 30, 36]
    header_row(ws, 1, headers, widths)
    for r, t in enumerate(technos, start=2):
        vals = [
            t.get("techno", ""), t.get("citee_ou", ""),
            t.get("etat_de_lart", ""), t.get("maturite", ""),
            t.get("alternatives", ""), t.get("vigilance", ""),
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
    ws.auto_filter.ref = f"A1:F{len(technos) + 1}"
    ws.freeze_panes = "A2"


def build_go_no_go(wb, data):
    items = data.get("go_no_go")
    if not items:
        return
    ws = wb.create_sheet("Go-No-go")
    headers = ["Critère", "Constat", "Poids dans la décision"]
    widths = [34, 60, 20]
    header_row(ws, 1, headers, widths)
    for r, g in enumerate(items, start=2):
        vals = [g.get("critere", ""), g.get("constat", ""), g.get("poids", "")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
    ws.auto_filter.ref = f"A1:C{len(items) + 1}"
    ws.freeze_panes = "A2"


def build_questions_acheteur(wb, data):
    questions = data.get("questions_acheteur")
    if not questions:
        return
    ws = wb.create_sheet("Questions à l'acheteur")
    headers = ["Question", "Article CCTP concerné", "Justification"]
    widths = [55, 26, 55]
    header_row(ws, 1, headers, widths)
    for r, q in enumerate(questions, start=2):
        vals = [q.get("question", ""), q.get("article", ""),
                q.get("justification", "")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
    ws.auto_filter.ref = f"A1:C{len(questions) + 1}"
    ws.freeze_panes = "A2"


def main():
    if len(sys.argv) != 3:
        sys.exit("Usage: python generate_exigences_xlsx.py input.json output.xlsx")
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)
    wb = Workbook()
    build_exigences(wb, data)
    build_jalons(wb, data)
    build_synthese_texte(wb, "Entité émettrice", data.get("entite_emettrice"))
    build_personnes(wb, data)
    build_synthese_texte(wb, "Secteur & industrie", data.get("secteur_industrie"))
    build_technologies(wb, data)
    build_go_no_go(wb, data)
    build_questions_acheteur(wb, data)
    build_texte_libre(wb, "Format de réponse imposé", data.get("format_reponse_impose"))
    wb.save(sys.argv[2])
    print(f"OK: {sys.argv[2]}")


if __name__ == "__main__":
    main()
